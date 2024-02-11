import openpyxl
import os

def copy_sentences(input_file, output_file):
    # Load the input workbook
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input.active

    # Create a new workbook for output
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active

    # Dictionary to hold sentences for each code
    code_sentences = {}

    # Iterate through rows and populate the dictionary
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        code = int(row[1]) if row[1] is not None else None  # Convert to integer if not None
        if code is not None:
            sentence = row[2]  # 'Sentence' column
            if code not in code_sentences:
                code_sentences[code] = []
            code_sentences[code].append(sentence)

    # Write sentences to the output worksheet
    for code, sentences in code_sentences.items():
        if code is not None:
            col_letter = chr(ord('A') + code - 1)  # Convert code to corresponding column letter
            for i, sentence in enumerate(sentences, start=1):
                ws_output[f"{col_letter}{i}"] = sentence

    # Save the output workbook
    wb_output.save(output_file)


if __name__ == "__main__":

    # Example usage:
    input_file = input("Input file:")
    output_file = input_file.replace('.xlsx', '_output.xlsx')
    copy_sentences(input_file, output_file)
